"""Excel export utility methods."""
from typing import Any, Dict, Iterable, List, Optional, Set

from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from raesl.doc import lines
from raesl.excel.defaults import MONO, WRAP
from ragraph.graph import Graph
from ragraph.node import Node


def get_all_tags(nodes: List[Node]) -> List[str]:
    """Get all tagged comment keys from a list of nodes."""
    tags: List[str] = []
    seen: Set[str] = set()
    for node in nodes:
        node_tags = [
            tag for tag in node.annotations.esl_info["tagged_comments"].keys() if tag not in seen
        ]
        tags.extend(node_tags)
        seen.update(node_tags)
    return tags


def make_table(
    ws: Worksheet,
    name: str = "Table",
    min_row: int = 1,
    max_row: int = 1,
    min_col: int = 1,
    max_col: int = 1,
    style: TableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True),
) -> Table:
    """Make a table of a cell range in a worksheet."""
    ref = CellRange(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)
    table = Table(name=name, displayName=name, ref=ref.coord)
    table.tableStyleInfo = style
    ws.add_table(table)
    return table


def apply_styling(ws: Worksheet, headers: List[str], defaults: Dict[str, Any] = dict()):
    """Apply styling to columns given some default option dictionary."""
    for i, header in enumerate(headers):
        idx = get_column_letter(i + 1)
        for cell in ws[idx]:
            styles = defaults.get(header, dict()).get("styles", dict())
            cell.font = styles.get("font", MONO)
            cell.alignment = styles.get("alignment", WRAP)
        width = defaults.get(header, dict()).get("width", 60)
        if width is not None:
            ws.column_dimensions[idx].width = width


def parent_component(requirement: Node, skip: Optional[str] = "world") -> str:
    """Get parent component name of a requirement."""
    path = ".".join(requirement.name.split(".")[:-1])
    return lines.node_path(path, arrows=False, skip=skip) if skip else path


def parent_def(graph: Graph, requirement: Node) -> str:
    """Get the parent (component) definition of a requirement."""
    parent_comp = ".".join(requirement.name.split(".")[:-1])
    parent_def = graph[parent_comp].annotations.esl_info["definition_name"]
    return parent_def


def format_multiline(comments: List[str]) -> str:
    """Format multiline (list) text."""
    return "\n".join(c.rstrip("\\") for c in comments)


def requirement_kind(requirement: Node) -> str:
    """Get requirement kind."""
    info = requirement.annotations.esl_info
    if requirement.kind == "function_spec":
        return info["sub_kind"]
    elif requirement.kind == "design_spec":
        return "design"
    elif requirement.kind == "behavior_spec":
        return "behavior"
    else:
        return requirement.kind


def dedupe(iterable: Iterable) -> List[Any]:
    """Deduplicate any iterable into a list where the first occurrence is preserved."""
    seen: Set[Any] = set()
    unique: List[Any] = list()
    for item in iterable:
        if item in seen:
            continue
        seen.add(item)
        unique.append(item)
    return unique
