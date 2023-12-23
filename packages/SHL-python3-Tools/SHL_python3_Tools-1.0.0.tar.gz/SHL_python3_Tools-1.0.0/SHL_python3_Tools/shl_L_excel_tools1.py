import openpyxl
from openpyxl.styles import PatternFill, Font


class PyExceuteExcel:

    def __init__(self, read_excel_path, write_excel_path):
        self.read_excel_path = read_excel_path
        self.write_excel_path = write_excel_path

    def getRows(self):
        # 打开工作簿
        workbook = openpyxl.load_workbook(self.read_excel_path)

        # 选择第一个工作表
        sheet = workbook.active

        # 从第一个工作表中读取两列数据
        two_columns_data = []
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=2):
            row_data = [cell.value for cell in row]
            two_columns_data.append(row_data)
        # # 输出结果
        # for i in two_columns_data:
        #     print(i)
        # print("===================")
        # print(two_columns_data)
        # 将每一行数组转换为字典
        # dict_list = [dict(zip(['column1', 'column2'], row)) for row in two_columns_data]

        # 转换为字典
        result_dict = {item[0]: item[1] for item in two_columns_data}

        for key, value in result_dict.items():
            print(key, ":", value)
            # 输出转换后的字典列表
        print("===================")

        # 关闭工作簿
        workbook.close()

        return result_dict

    # 整体替换
    def setRows(self, data):
        # 创建一个新的工作簿
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # 将数组写入工作表
        for row_index, row in enumerate(data, start=1):
            for col_index, cell_value in enumerate(row, start=1):
                print(col_index, cell_value)
                sheet.cell(row=row_index, column=col_index, value=cell_value)

        # 保存工作簿
        workbook.save(self.write_excel_path)

    # 追加
    def appendRows(self, data):
        # 加载现有的工作簿
        workbook = openpyxl.load_workbook(self.write_excel_path)
        sheet = workbook.active

        # 确定追加数据的起始行
        start_row = sheet.max_row + 1

        # 将数组写入工作表
        for row_index, row in enumerate(data, start=start_row):
            for col_index, cell_value in enumerate(row, start=1):
                sheet.cell(row=row_index, column=col_index, value=cell_value)

        # 保存工作簿
        workbook.save(self.write_excel_path)

    # 追加+文本格式
    def appendRowsWithTextFormat(self, data):
        # 加载现有的工作簿
        workbook = openpyxl.load_workbook(self.write_excel_path)
        sheet = workbook.active

        # 确定追加数据的起始行
        start_row = sheet.max_row + 1

        # 创建文本格式样式
        text_format = Font(name='Arial', size=12, bold=False, color='000000')

        # 将数组写入工作表
        for row_index, row in enumerate(data, start=start_row):
            for col_index, cell_value in enumerate(row, start=1):
                # 将数据转换为文本格式并写入单元格
                # if type(cell_value) == str and "=" in cell_value:
                #     cell_value = cell_value.replace("=", " =")
                # cell_value_new = " "+str(cell_value)
                cell_value_new = "\t"+str(cell_value)
                sheet.cell(row=row_index, column=col_index, value=str(cell_value_new))
                # 设置单元格样式为文本格式
                sheet.cell(row=row_index, column=col_index).number_format = '@'  # 设置格式为纯文本，@ 表示文本格式
                sheet.cell(row=row_index, column=col_index).font = text_format

        # 保存工作簿
        workbook.save(self.write_excel_path)
