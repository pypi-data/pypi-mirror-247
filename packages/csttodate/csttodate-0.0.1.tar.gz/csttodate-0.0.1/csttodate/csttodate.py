

def read_excel(path,col,save_path):
    import openpyxl
    xlsx_path = path
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.worksheets[0] #获取第一页sheet
    cols = ws.max_column #获取有效的数据的最大列数
    rows = ws.max_row  #最大行数
    for row in range(2,rows+1):
        old_value=ws.cell(row=row, column=col).value  # 获取(row,col)单元格的值
        new_value=cst_to_date(old_value)
        ws.cell(row=row, column=col).value = new_value  # 将(row,col)单元格的值 变成一个[1,5]的随机整数
    # 此处有一个注意事项
    # cell(r,c)中的行r和列c 是从1开始。而不是0。
    # wb = openpyxl.load_workbook(xlsx_path)
    wb.save(save_path)  # 设置保存的路径 和保存的文件名：2.xlsx



def cst_to_date(cst_time):
    date_str=cst_time.split(' ')
    if date_str[1]=='Nov':
        date_str[1]='11'
    elif date_str[1]=='Dec':
        date_str[1]='12'
    elif date_str[1]=='Oct':
        date_str[1]='10'
    elif date_str[1]=='Sep':
        date_str[1] = '09'
    elif date_str[1] == 'Aug':
        date_str[1] = '08'
    elif date_str[1] == 'Jul':
        date_str[1] = '07'
    elif date_str[1] == 'Jun':
        date_str[1] = '06'
    elif date_str[1] == 'May':
        date_str[1] = '05'
    elif date_str[1] == 'Apr':
        date_str[1] = '04'
    elif date_str[1] == 'Mar':
        date_str[1] = '03'
    elif date_str[1] == 'Feb':
        date_str[1] = '02'
    elif date_str[1] == 'Jan':
        date_str[1] = '01'
    date_res=date_str[-1]+'-'+date_str[1]+'-'+date_str[2]+' '+date_str[3]
    return date_res



